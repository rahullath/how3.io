import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, List, Optional, Tuple, Any
import os
import sys

def load_excel(file_path: str) -> pd.DataFrame:
    """
    Load an Excel file with multi-level headers and fix unnamed headers.

    Args:
        file_path: Path to the Excel file.

    Returns:
        A pandas DataFrame with flattened and fixed column names.
    """
    # Load the Excel file with multi-level headers
    df = pd.read_excel(file_path, header=[0, 1])

    # Fix unnamed headers explicitly
    new_columns = []
    for top, bottom in df.columns:
        if 'Unnamed' in top:
            new_columns.append(bottom.strip())
        else:
            new_columns.append(f"{top.strip()}_{bottom.strip()}")
    df.columns = new_columns

    return df

class EarningsQualityAnalyzer:
    """
    A specialized class for analyzing the quality and sustainability of earnings/revenue
    for cryptocurrency projects.
    """
    
    def __init__(self, df: pd.DataFrame):
        """Initialize with a dataframe containing crypto project metrics."""
        self.df = df
        self._fix_column_names()
        self.revenue_columns = self._find_revenue_columns()
        self.stability_columns = self._find_stability_columns()
        self.diversification_columns = self._find_diversification_columns()
    
    def _fix_column_names(self):
        """Fix common issues with column names."""
        # Check if we have Project and Market sector columns
        has_project = False
        has_market_sector = False
        
        # Look for Project column
        for col in self.df.columns:
            if col == 'Project' or (isinstance(col, str) and col.lower() == 'project'):
                has_project = True
                if col != 'Project':
                    self.df.rename(columns={col: 'Project'}, inplace=True)
                    print(f"Renamed '{col}' to 'Project'")
                break
        
        # Look for Market sector column
        for col in self.df.columns:
            if col == 'Market sector' or (isinstance(col, str) and col.lower() in ['market sector', 'marketsector', 'sector']):
                has_market_sector = True
                if col != 'Market sector':
                    self.df.rename(columns={col: 'Market sector'}, inplace=True)
                    print(f"Renamed '{col}' to 'Market sector'")
                break
        
        # If we still don't have Project column, look for unnamed columns
        if not has_project or not has_market_sector:
            # Check if there are unnamed columns that might contain our data
            unnamed_cols = [col for col in self.df.columns if 'Unnamed:' in str(col)]
            
            if len(unnamed_cols) >= 2:
                print(f"Found {len(unnamed_cols)} unnamed columns. Checking if they contain project and sector data...")
                
                # Check if the first row might contain headers
                if len(self.df) > 0:
                    first_row = self.df.iloc[0]
                    found_project = False
                    found_sector = False
                    
                    for i, col in enumerate(unnamed_cols):
                        val = first_row[col]
                        if isinstance(val, str):
                            if val.lower() in ['project', 'name', 'token', 'cryptocurrency']:
                                # This might be the project column
                                self.df.rename(columns={col: 'Project'}, inplace=True)
                                print(f"Renamed '{col}' to 'Project' based on first row value")
                                found_project = True
                            elif 'sector' in val.lower() or 'category' in val.lower():
                                # This might be the market sector column
                                self.df.rename(columns={col: 'Market sector'}, inplace=True)
                                print(f"Renamed '{col}' to 'Market sector' based on first row value")
                                found_sector = True
                    
                    # If we found headers in first row, drop it
                    if found_project or found_sector:
                        print("First row appears to contain headers. Dropping it...")
                        self.df = self.df.iloc[1:].reset_index(drop=True)
                
                # If we still don't have Project column, use the first unnamed column
                if 'Project' not in self.df.columns and len(unnamed_cols) > 0:
                    self.df.rename(columns={unnamed_cols[0]: 'Project'}, inplace=True)
                    print(f"Using first unnamed column '{unnamed_cols[0]}' as 'Project'")
                
                # If we still don't have Market sector column, use the second unnamed column
                if 'Market sector' not in self.df.columns and len(unnamed_cols) > 1:
                    self.df.rename(columns={unnamed_cols[1]: 'Market sector'}, inplace=True)
                    print(f"Using second unnamed column '{unnamed_cols[1]}' as 'Market sector'")
        
        # Check if we have data in these columns
        if 'Project' in self.df.columns:
            # If all values are null, try to find another column
            if self.df['Project'].isna().all():
                print("Project column contains all null values. Looking for alternative...")
                
                # Try to find a column with project names
                for col in self.df.columns:
                    if col != 'Project' and col != 'Market sector':
                        non_null_vals = self.df[col].dropna()
                        if len(non_null_vals) > 0 and all(isinstance(x, str) for x in non_null_vals):
                            self.df['Project'] = self.df[col]
                            print(f"Using '{col}' values for Project")
                            break
        
        # Same for Market sector
        if 'Market sector' in self.df.columns:
            # If all values are null, try to find another column
            if self.df['Market sector'].isna().all():
                print("Market sector column contains all null values. Looking for alternative...")
                
                # Try to find a column with sector names
                for col in self.df.columns:
                    if col != 'Project' and col != 'Market sector':
                        non_null_vals = self.df[col].dropna()
                        if len(non_null_vals) > 0 and all(isinstance(x, str) for x in non_null_vals):
                            if any('blockchain' in str(x).lower() or 'defi' in str(x).lower() for x in non_null_vals):
                                self.df['Market sector'] = self.df[col]
                                print(f"Using '{col}' values for Market sector")
                                break
        
        # Print the results
        if 'Project' in self.df.columns:
            non_null_projects = self.df['Project'].dropna()
            print(f"Project column has {len(non_null_projects)} non-null values")
            if len(non_null_projects) > 0:
                print(f"Sample projects: {non_null_projects.iloc[:5].tolist()}")
        
        if 'Market sector' in self.df.columns:
            non_null_sectors = self.df['Market sector'].dropna()
            print(f"Market sector column has {len(non_null_sectors)} non-null values")
            if len(non_null_sectors) > 0:
                print(f"Sample sectors: {non_null_sectors.unique()[:5].tolist()}")
    
    def _find_revenue_columns(self) -> Dict[str, List[str]]:
        """
        Find columns specifically for revenue stability calculations.

        Returns:
            Dict with explicitly defined stability columns.
        """
        revenue_cols = {
            'stability': []
        }

        # Define stability-specific keywords and periods
        stability_keywords = ['fees', 'supply-side fees', 'earnings']
        stability_periods = ['30d trend', '90d trend', '180d trend', '365d trend']

        for col in self.df.columns:
            col_lower = col.lower()
            if any(keyword in col_lower for keyword in stability_keywords) \
                    and any(period in col_lower for period in stability_periods):
                revenue_cols['stability'].append(col)

        # Print found stability columns for debugging
        print(f"Found {len(revenue_cols['stability'])} stability columns")
        if revenue_cols['stability']:
            print(f"Sample stability columns: {revenue_cols['stability'][:3]}")

        return revenue_cols
    
    def _get_time_period_order(self, col_name: str) -> int:
        """
        Return a sortable value based on the time period in the column name.

        Args:
            col_name: The name of the column.

        Returns:
            An integer representing the order of the time period.
        """
        # Define a priority for sorting columns
        periods = {
            '24h': 1,
            '7d': 2,
            '30d': 3,
            '90d': 4,
            '180d': 5,
            '365d': 6
        }
        for period, order in periods.items():
            if period in col_name:
                return order
        # Default high number if no known period is found
        return 999
    
    def _find_stability_columns(self) -> List[str]:
        """Find columns that can be used to measure revenue stability."""
        stability_cols = []
        
        # Look for columns with change/trend information
        stability_keywords = [
            'change', 'trend', 'growth', 'volatility', 'stability'
        ]
        
        time_periods = ['30d', '90d', '180d', '365d']
        
        # Find revenue/fee columns with change metrics
        for col in self.df.columns:
            col_str = str(col).lower()
            
            # Check if column contains revenue/fee and change/trend
            if any(rev in col_str for rev in ['revenue', 'fees', 'earnings']):
                if any(stab in col_str for stab in stability_keywords):
                    if any(period in col_str for period in time_periods):
                        stability_cols.append(col)
        
        print(f"Found {len(stability_cols)} stability columns")
        if stability_cols:
            print(f"Sample columns: {stability_cols[:3]}")
        
        return stability_cols
    
    def _find_diversification_columns(self) -> List[str]:
        """Find columns related to revenue diversification."""
        diversification_cols = []
        
        # Look for columns with revenue breakdown information
        diversification_keywords = [
            'breakdown', 'source', 'category', 'segment', 'diversification',
            'cost of revenue', 'operating expenses', 'gross profit'
        ]
        
        for col in self.df.columns:
            col_str = str(col).lower()
            
            # Check for revenue diversification metrics
            if any(rev in col_str for rev in ['revenue', 'fees', 'earnings']):
                if any(div in col_str for div in diversification_keywords):
                    diversification_cols.append(col)
        
        print(f"Found {len(diversification_cols)} diversification columns")
        if diversification_cols:
            print(f"Sample columns: {diversification_cols[:3]}")
        
        return diversification_cols
    
    def calculate_quarterly_values(self, row: pd.Series) -> Dict[str, float]:
        """Calculate quarterly revenue/fee values for a project."""
        quarterly_values = {}

        # Try to find best quarterly data
        quarterly_cols = [col for col in self.revenue_columns['primary'] if '90d' in str(col).lower()]

        if quarterly_cols:
            for i, col in enumerate(quarterly_cols[:4]):  # Use up to 4 quarters
                value = self._get_numeric_value(row, col)
                if value is not None:
                    quarterly_values[f'Q{i+1}'] = value

        # If we don't have 4 quarters, try to derive from yearly data
        if len(quarterly_values) < 4:
            yearly_cols = [col for col in self.revenue_columns['primary'] if '365d' in str(col).lower()]
            if yearly_cols:
                yearly_col = yearly_cols[0]
                yearly_value = self._get_numeric_value(row, yearly_col)
                if yearly_value is not None:
                    missing_quarters = 4 - len(quarterly_values)
                    for i in range(missing_quarters):
                        q_num = len(quarterly_values) + 1
                        quarterly_values[f'Q{q_num}'] = yearly_value / 4

        return quarterly_values
    
    def _get_numeric_value(self, row: pd.Series, col_name: str) -> Optional[float]:
        """Extract a numeric value from the row."""
        value = row.get(col_name, np.nan)

        if pd.isna(value):
            return None

        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            try:
                cleaned = ''.join(c for c in value if c.isdigit() or c in '.-')
                return float(cleaned) if cleaned else None
            except (ValueError, TypeError):
                return None

        return None
    
    def calculate_stability_score(self, row: pd.Series) -> Tuple[Optional[float], Dict[str, Any]]:
        """
        Calculate revenue stability score (0-100) based on prioritized metrics.

        Args:
            row: A pandas Series representing a single row of the DataFrame.

        Returns:
            Tuple of (stability_score, explanation_dict).
        """
        # Define priority metrics
        priority_metrics = [
            ("Revenue", ["Revenue_7d sum", "Revenue_30d sum"], ["Revenue_30d trend", "Revenue_90d trend"]),
            ("Earnings", ["Earnings_7d sum", "Earnings_30d sum"], ["Earnings_30d trend", "Earnings_90d trend"]),
            ("Supply-side fees", ["Supply-side fees_7d sum", "Supply-side fees_30d sum"], ["Supply-side fees_30d trend", "Supply-side fees_90d trend"]),
            ("Fees", ["Fees_7d sum", "Fees_30d sum"], ["Fees_30d trend", "Fees_90d trend"]),
        ]

        explanation = {
            "chosen_metric": None,
            "metric_values": {},
            "trends": {},
            "stability_score": None,
            "method": "No sufficient data"
        }

        # Iterate through priority metrics
        for metric_name, sums, trends in priority_metrics:
            # Check if sum and trend data are available
            sums_available = all(pd.notna(row.get(col)) and row.get(col, 0) > 0 for col in sums)
            trends_available = all(pd.notna(row.get(trend)) for trend in trends)

            if sums_available and trends_available:
                # Use this metric for stability calculation
                explanation["chosen_metric"] = metric_name
                explanation["metric_values"] = {col: row.get(col) for col in sums}
                explanation["trends"] = {trend: row.get(trend) for trend in trends}

                # Calculate stability score from trends
                trend_values = np.array([abs(row.get(trend)) for trend in trends])
                avg_trend = np.mean(trend_values)

                # Stability is inversely proportional to trend volatility
                stability_score = max(0, 100 - (avg_trend * 100))
                explanation["stability_score"] = round(stability_score, 2)
                explanation["method"] = "Calculated from trends and sums"
                return stability_score, explanation

        # If no metric is fully available, return NaN
        return np.nan, explanation
    
    def calculate_revenue_diversification(self, row: pd.Series, cols: List[str]) -> Tuple[Optional[float], Dict[str, Any]]:
        """
        Calculate revenue diversification score (0-100) based on distribution of revenue sources.

        Args:
            row: A pandas Series representing a single row of the DataFrame.
            cols: List of diversification-related columns to analyze.

        Returns:
            Tuple of (diversification_score, explanation_dict).
        """
        explanation = {
            'diversification_values': {},
            'diversification_score': None,
            'method_used': 'No valid data'
        }

        # Handle missing data explicitly
        if row[cols].isnull().all():
            explanation['method_used'] = 'No diversification data available'
            return np.nan, explanation

        # Calculate standard deviation and mean
        diversification_std = row[cols].std()
        diversification_mean = row[cols].mean()

        # Handle zero mean explicitly
        if diversification_mean <= 0:
            explanation['method_used'] = 'No diversification (mean is zero)'
            return 0, explanation

        # Diversification score scales with standard deviation
        diversification_score = 100 * (diversification_std / diversification_mean)
        diversification_score = max(0, min(100, diversification_score))
        explanation['diversification_values'] = row[cols].to_dict()
        explanation['diversification_score'] = round(diversification_score, 2)
        explanation['method_used'] = 'Calculated from diversification data'

        return round(diversification_score, 2), explanation
    
    def calculate_earnings_quality(self, row: pd.Series, min_projects_for_percentile: int = 3) -> Dict[str, Any]:
        """
        Calculate overall earnings quality score combining stability and diversification.

        Args:
            row: A pandas Series representing a single row of the DataFrame.
            min_projects_for_percentile: Minimum number of projects needed in sector for percentile calculations.

        Returns:
            Dictionary with earnings quality scores and explanations.
        """
        project_val = row.get('Project', "Unknown")
        market_sector = row.get('Market sector', "Unknown")

        result = {
            'stability': {
                'score': None,
                'percentile': None,
                'explanation': None
            },
            'diversification': {
                'score': None,
                'percentile': None,
                'explanation': None
            },
            'overall_score': None,
            'metrics_used': [],
            'complete_score': False
        }

        # Stability calculation
        stability_score, stability_explanation = self.calculate_stability_score(row)
        result['stability']['score'] = stability_score
        result['stability']['explanation'] = stability_explanation

        if stability_score is not None:
            result['metrics_used'].append('stability')

        # Diversification calculation
        diversification_score, diversification_explanation = self.calculate_revenue_diversification(row, self.diversification_columns)
        result['diversification']['score'] = diversification_score
        result['diversification']['explanation'] = diversification_explanation

        if diversification_score is not None:
            result['metrics_used'].append('diversification')

        # Calculate overall earnings quality score
        valid_percentiles = []
        if stability_score is not None:
            valid_percentiles.append(stability_score)
        if diversification_score is not None:
            valid_percentiles.append(diversification_score)

        if valid_percentiles:
            # Weight stability higher than diversification if both are available
            if len(valid_percentiles) == 2:
                result['overall_score'] = 0.7 * stability_score + 0.3 * diversification_score
                result['complete_score'] = True
            else:
                result['overall_score'] = valid_percentiles[0]
                result['complete_score'] = len(valid_percentiles) == 2

        return result


def calculate_revenue_magnitude(row: pd.Series, max_revenue_sum: float) -> float:
    """
    Calculate the revenue magnitude score (0-100) based on the project's revenue sum.

    Args:
        row: A pandas Series representing a single row of the DataFrame.
        max_revenue_sum: The maximum revenue sum across all projects.

    Returns:
        Revenue Magnitude Score (0-100).
    """
    revenue_sum = row.get('Revenue_30d sum', 0)  # Use 30d sum as the primary revenue metric
    if revenue_sum <= 0:
        return 0  # No revenue, no magnitude score

    # Normalize using log scale
    magnitude_score = 100 * np.log(revenue_sum + 1) / np.log(max_revenue_sum + 1)
    return round(magnitude_score, 2)


def calculate_revenue_quality(row: pd.Series, max_revenue_sum: float) -> Tuple[Optional[float], Dict[str, Any]]:
    """
    Calculate the combined revenue quality score (0-100) based on stability and magnitude.

    Args:
        row: A pandas Series representing a single row of the DataFrame.
        max_revenue_sum: The maximum revenue sum across all projects.

    Returns:
        Tuple of (Revenue Quality Score, Explanation Dictionary).
    """
    # Calculate Stability Score
    analyzer = EarningsQualityAnalyzer(df)  # Create an instance of the class
    stability_score, stability_explanation = analyzer.calculate_stability_score(row)

    # Calculate Magnitude Score
    magnitude_score = calculate_revenue_magnitude(row, max_revenue_sum)

    # Combine Stability and Magnitude Scores
    if stability_score is not None and magnitude_score > 0:
        revenue_quality_score = (stability_score * magnitude_score) / 100
    else:
        revenue_quality_score = 0  # If either score is missing or zero, quality is zero

    # Explanation
    explanation = {
        'stability_score': stability_score,
        'magnitude_score': magnitude_score,
        'revenue_quality_score': round(revenue_quality_score, 2),
        'method': 'Combined magnitude (log-scaled) and stability (trend-based)'
    }

    return revenue_quality_score, explanation


def enhance_earnings_quality_analysis(df: pd.DataFrame) -> pd.DataFrame:
    """
    Enhance earnings quality analysis for all projects in the dataset.

    Args:
        df: DataFrame with crypto project metrics.

    Returns:
        DataFrame with enhanced earnings quality scores.
    """
    # Find the maximum revenue sum for normalization
    max_revenue_sum = df['Revenue_30d sum'].max()

    # Create result columns
    results = []

    # Process each project
    for idx, row in df.iterrows():
        project_name = row.get('Project', "Unknown")
        market_sector = row.get('Market sector', "Unknown")

        if pd.isna(project_name) or project_name.lower() in ['project', 'name', '#', 'token', 'unknown']:
            continue

        print(f"Processing {project_name} ({market_sector})...")

        # Calculate Revenue Quality Score
        revenue_quality_score, quality_explanation = calculate_revenue_quality(row, max_revenue_sum)

        # Add results to the output
        result = {
            'Project': project_name,
            'Market Sector': market_sector,
            'Revenue Quality Score': revenue_quality_score,
            'Quality Explanation': quality_explanation
        }
        results.append(result)

    # Convert to DataFrame
    results_df = pd.DataFrame(results)
    return results_df


def visualize_earnings_quality(results_df: pd.DataFrame, top_n: int = 10) -> None:
    """
    Create visualizations for earnings quality analysis.

    Args:
        results_df: DataFrame with earnings quality results.
        top_n: Number of top projects to show in visualizations.
    """
    import seaborn as sns
    import matplotlib.pyplot as plt

    # Check if required columns exist
    required_cols = ["Revenue Quality Score"]
    for col in required_cols:
        if col not in results_df.columns:
            print(f"Column '{col}' is missing from results. Visualization skipped.")
            return

    # Drop rows with NaNs in key columns
    clean_results = results_df.dropna(subset=required_cols)

    if clean_results.empty:
        print("No valid data points available after cleaning. Visualization skipped.")
        return

    # Bar plot: Top projects by Revenue Quality Score
    top_results = clean_results.nlargest(top_n, "Revenue Quality Score")
    plt.figure(figsize=(12, 8))
    sns.barplot(
        data=top_results,
        x="Revenue Quality Score",
        y="Project",
        palette="viridis"
    )

    plt.title("Top Projects by Revenue Quality Score")
    plt.xlabel("Revenue Quality Score")
    plt.ylabel("Project")
    plt.tight_layout()
    plt.savefig("revenue_quality_scores.png")
    plt.show()


def debug_data_structure(file_path: str) -> None:
    """Debug the data structure of the file to assist with parsing."""
    print(f"\nDEBUGGING FILE: {file_path}")
    
    try:
        # Determine file type
        if file_path.lower().endswith('.csv'):
            # For CSV, read the first few rows and report columns
            df = pd.read_csv(file_path, nrows=5)
            print(f"CSV file with {df.shape[1]} columns")
            print(f"Column names: {df.columns.tolist()[:10]}...")
            
            # Check for unnamed columns
            unnamed = [col for col in df.columns if 'Unnamed:' in str(col)]
            if unnamed:
                print(f"Found {len(unnamed)} unnamed columns")
                
                # Check first row as potential headers
                first_row = df.iloc[0]
                print(f"First row values: {first_row.tolist()[:10]}...")
                
        elif file_path.lower().endswith(('.xlsx', '.xls')):
            # For Excel, check sheet names and first few rows
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            print(f"Excel file with sheets: {wb.sheetnames}")
            
            # Get first sheet
            sheet = wb[wb.sheetnames[0]]
            print(f"First sheet: {sheet.title}")
            
            # Report dimensions
            print(f"Sheet dimensions: {sheet.max_row} rows, {sheet.max_column} columns")
            
            # Check first few cells
            first_row = []
            for cell in list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]:
                first_row.append(cell)
            print(f"First row values: {first_row[:10]}...")
            
            # Check second row
            second_row = []
            for cell in list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]:
                second_row.append(cell)
            print(f"Second row values: {second_row[:10]}...")
            
        else:
            print(f"Unsupported file format: {file_path}")
    
    except Exception as e:
        print(f"Error debugging file: {e}")


# Example usage
file_path = "tt-master-data_2025-03-27.xlsx"
df = load_excel(file_path)

# Verify the flattened column names
print("Flattened Columns:")
print(df.columns.tolist())

# Verify column names in the CSV
output_file = 'earnings_quality_results.csv'
df_results = pd.read_csv(output_file)

print("Columns in the CSV file:")
print(df_results.columns.tolist())

if __name__ == "__main__":
    import sys
    import os
    
    # Check for debugging flag
    debug_mode = '--debug' in sys.argv
    if debug_mode:
        sys.argv.remove('--debug')
    
    # Find input file
    data_file = None
    if len(sys.argv) > 1:
        data_file = sys.argv[1]
    else:
        # Try common filenames
        for filename in ['tt-master-data_2025-03-27.xlsx', 'ttmasterdata_20250327.xlsx', 'crypto_data.xlsx']:
            if os.path.exists(filename):
                data_file = filename
                break
    
    if not data_file or not os.path.exists(data_file):
        print("Error: Data file not found")
        print("Usage: python earnings_quality.py [data_file.xlsx] [--debug]")
        sys.exit(1)
    
    # Run in debug mode if requested
    if debug_mode:
        debug_data_structure(data_file)
        sys.exit(0)
    
    print(f"Loading data from {data_file}...")
    
    # Load data
    try:
        if data_file.endswith('.csv'):
            df = pd.read_csv(data_file)
        else:
            df = load_excel(data_file)
        
        print(f"Loaded data with {len(df)} rows and {df.shape[1]} columns")
    except Exception as e:
        print(f"Error loading data: {e}")
        sys.exit(1)
    
    # Filter projects by market cap
    MIN_MARKET_CAP = 115_000_000  # 115 million

    if ('Market cap (circulating)', 'Latest') in df.columns:
        df_filtered = df[df[('Market cap (circulating)', 'Latest')] >= MIN_MARKET_CAP]
        print(f"Filtered data to {len(df_filtered)} projects with market cap >= {MIN_MARKET_CAP}")
    else:
        print("Market cap column not found. Proceeding without filtering.")
        df_filtered = df

    # Run analysis on filtered data
    results = enhance_earnings_quality_analysis(df_filtered)
    
    # Check if we got results
    if results.empty:
        print("No valid results were generated. Try running with --debug to diagnose the issue.")
        sys.exit(1)
    
    # Save results
    output_file = 'earnings_quality_results.csv'
    results.to_csv(output_file, index=False)
    print(f"Results saved to {output_file}")
    
    # Create visualizations
    visualize_earnings_quality(results)
    print("Visualizations saved to revenue_quality_scores.png")